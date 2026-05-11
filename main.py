import sys, os, re, csv, time, glob, datetime as dt, tkinter as tk
from tkinter import ttk, messagebox, filedialog
import numpy as np, matplotlib.pyplot as plt
plt.switch_backend('TkAgg')
def _p(s):
    m=re.match(r'^(\d{4})-(\d{2})$',s)
    if m: return (int(m.group(1)),int(m.group(2)))
    try: return float(s)
    except: return None
def _pr(raw):
    if any(isinstance(r[0],tuple) for r in raw):
        d=[r[0] if isinstance(r[0],tuple) else (int(r[0]) if r[0]==int(r[0]) else int(r[0]),1) for r in raw]
        md=min(d,key=lambda x:(x[0],x[1])); xv,yv,lb=[],[],[]
        for xr,yr in raw:
            if isinstance(xr,tuple):
                y,m=xr; mo=(y-md[0])*12+(m-md[1]); lb.append(f"{y}-{m:02d}")
            else:
                y=int(xr) if xr==int(xr) else xr; mo=(y-md[0])*12+(1-md[1]); lb.append(str(y))
            xv.append(mo); yv.append(yr)
        return np.array(xv,dtype=float),np.array(yv,dtype=float),lb
    else:
        xv=[float(r[0]) for r in raw]; yv=[r[1] for r in raw]
        lb=[str(int(v)) if v==int(v) else str(v) for v in xv]
        return np.array(xv,dtype=float),np.array(yv,dtype=float),lb
def lcsv(fp):
    raw=[]
    if os.path.splitext(fp)[1]=='.csv':
        with open(fp,encoding='utf-8') as f:
            s=f.read(2048); f.seek(0)
            try: dl=csv.Sniffer().sniff(s,delimiters=[',',';','\t']).delimiter
            except: dl=',' if ',' in s else (';' if ';' in s else ('\t' if '\t' in s else ','))
            for r in csv.reader(f,delimiter=dl):
                if len(r)<2: continue
                x=_p(r[0].strip())
                if x is None: continue
                try: raw.append((x,float(r[1])))
                except: pass
    else:
        data=None
        for d in [',',';',None]:
            try: data=np.loadtxt(fp,delimiter=d,comments='#')
            except: continue
            if data.ndim==2 and data.shape[1]==2: break
        if data is not None:
            x=data[:,0]; y=data[:,1]
            return np.array(x),np.array(y),[str(int(v)) if v==int(v) else str(v) for v in x]
        xv,yv=[],[]
        with open(fp) as f:
            for line in f:
                line=line.strip()
                if not line or line.startswith('#'): continue
                p=line.replace(',',' ').split()
                if len(p)>=2:
                    try: xv.append(float(p[0])); yv.append(float(p[1]))
                    except: pass
        if not xv: raise ValueError("Нет данных")
        x=np.array(xv); y=np.array(yv)
        return x,y,[str(int(v)) if v==int(v) else str(v) for v in x]
    if not raw: raise ValueError("Нет данных")
    return _pr(raw)
def _sd(s,dm):
    base=dt.datetime(1899,12,30) if dm==0 else dt.datetime(1904,1,1)
    try: return base+dt.timedelta(days=int(s))
    except: return None
def lex(fp):
    ext=os.path.splitext(fp)[1]; raw=[]
    if ext=='.xls':
        import xlrd; wb=xlrd.open_workbook(fp); sh=wb.sheet_by_index(0); dm=wb.datemode
        for r in range(sh.nrows):
            if sh.ncols<2: continue
            try: yv=float(str(sh.cell(r,1).value).replace(',','.'))
            except: continue
            if yv<=0: continue
            xr=None; c=sh.cell(r,0)
            if c.ctype==xlrd.XL_CELL_DATE:
                try:
                    t=xlrd.xldate_as_tuple(c.value,dm); yr,m,d=int(t[0]),int(t[1]),int(t[2])
                    xr=yr if (m==1 and d==1) else (yr,m)
                except: pass
            elif c.ctype==xlrd.XL_CELL_NUMBER:
                v=float(c.value)
                if 100<=v<=2100 and abs(v-round(v))<1e-9: xr=int(round(v))
                else:
                    d=_sd(c.value,dm)
                    if d: xr=d.year if (d.month==1 and d.day==1) else (d.year,d.month)
            elif c.ctype==xlrd.XL_CELL_TEXT:
                txt=str(c.value).strip(); m=re.match(r'^(\d{4})-(\d{2})$',txt)
                if m: xr=(int(m.group(1)),int(m.group(2)))
            if xr is not None: raw.append((xr,yv))
    else:
        import openpyxl; wb=openpyxl.load_workbook(fp,data_only=True); ws=wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=1,values_only=False):
            if not row or len(row)<2: continue
            try: yv=float(str(row[1].value).replace(',','.'))
            except: continue
            if yv<=0: continue
            xr=None; cell=row[0].value
            if isinstance(cell,dt.datetime):
                xr=cell.year if (cell.month==1 and cell.day==1) else (cell.year,cell.month)
            elif isinstance(cell,(int,float)):
                v=float(str(cell).replace(',','.'))
                if 100<=v<=2100 and abs(v-round(v))<1e-9: xr=int(round(v))
                else:
                    d=_sd(v,0)
                    if d: xr=d.year if (d.month==1 and d.day==1) else (d.year,d.month)
            elif isinstance(cell,str):
                txt=cell.strip(); m=re.match(r'^(\d{4})-(\d{2})$',txt)
                if m: xr=(int(m.group(1)),int(m.group(2)))
            if xr is not None: raw.append((xr,yv))
    if not raw: raise ValueError("Нет данных")
    return _pr(raw)
def _sm(w):
    try: w.state('zoomed')
    except:
        try: w.attributes('-zoomed',True)
        except: w.attributes('-fullscreen',True)
def m_inp():
    w=tk.Toplevel(); w.geometry('600x500'); w.title("Ручной ввод")
    mf=ttk.Frame(w,padding=20); mf.pack(fill=tk.BOTH,expand=True)
    ttk.Label(mf,text="Введите пары X и Y (мин. 2)",font=("Times New Roman",18,"bold")).pack(pady=10)
    cf=ttk.Frame(mf); cf.pack(fill=tk.BOTH,expand=True,padx=20,pady=10)
    cv=tk.Canvas(cf,highlightthickness=0); sb=ttk.Scrollbar(cf,orient=tk.VERTICAL,command=cv.yview)
    sf=ttk.Frame(cv); sf.bind("<Configure>",lambda e:cv.configure(scrollregion=cv.bbox("all")))
    cv.create_window((0,0),window=sf,anchor="nw"); cv.configure(yscrollcommand=sb.set)
    cv.pack(side=tk.LEFT,fill=tk.BOTH,expand=True); sb.pack(side=tk.RIGHT,fill=tk.Y)
    ent=[]
    def ar():
        r=len(ent)
        ttk.Label(sf,text=f"X{r+1}",font=("Calibri",14)).grid(row=r,column=0,padx=5,pady=2,sticky=tk.E)
        ex=ttk.Entry(sf,font=("Calibri",14),width=20); ex.grid(row=r,column=1,padx=5,pady=2)
        ttk.Label(sf,text=f"Y{r+1}",font=("Calibri",14)).grid(row=r,column=2,padx=5,pady=2,sticky=tk.E)
        ey=ttk.Entry(sf,font=("Calibri",14),width=20); ey.grid(row=r,column=3,padx=5,pady=2)
        ent.append((ex,ey))
    for _ in range(5): ar()
    bf=ttk.Frame(mf); bf.pack(pady=15)
    def ld():
        try:
            xv=[float(ex.get().replace(',','.')) for ex,_ in ent]
            yv=[float(ey.get().replace(',','.')) for _,ey in ent]
            if len(xv)<2: messagebox.showerror("Ошибка","Минимум 2 точки"); return
            lb=[str(int(v)) if v==int(v) else str(v) for v in xv]
            w.result=(np.array(xv),np.array(yv),lb); w.destroy()
        except: messagebox.showerror("Ошибка","Проверьте числа")
    ttk.Button(bf,text="Добавить строку",command=ar).pack(side=tk.LEFT,padx=10)
    ttk.Button(bf,text="Загрузить данные",command=ld).pack(side=tk.LEFT,padx=10)
    ttk.Button(bf,text="Отмена",command=lambda:(setattr(w,'result',None),w.destroy())).pack(side=tk.LEFT,padx=10)
    w.result=None; w.wait_window()
    if w.result is None: raise ValueError("Ввод отменён")
    return w.result
def pf(x,y,d): A=np.vander(x,d+1,increasing=True); return np.linalg.lstsq(A,y,rcond=None)[0]
def pv(coeffs,x): x=np.asarray(x); y=np.zeros_like(x,float); [y:=y+c*(x**i) for i,c in enumerate(coeffs)]; return y
def r2(yt,yp): return 1-np.sum((yt-yp)**2)/np.sum((yt-np.mean(yt))**2)
class App:
    def __init__(s,master):
        s.m=master; master.title("Аналитический инструмент"); _sm(master); master.configure(bg='#f0f0f0')
        style=ttk.Style(); style.configure("Big.TButton",font=("Calibri",16,"bold"),padding=15)
        c=ttk.Frame(master); c.place(relx=0.5,rely=0.5,anchor=tk.CENTER)
        ttk.Label(c,text="Добро пожаловать!",font=("Times New Roman",28,"bold")).pack(pady=20)
        ttk.Label(c,text="Выберите способ загрузки:",font=("Times New Roman",18)).pack(pady=10)
        ttk.Button(c,text="Загрузить из файлов (txt, CSV, Excel)",style="Big.TButton",command=s.lf).pack(pady=10)
        ttk.Button(c,text="Ввести вручную",style="Big.TButton",command=s.lm).pack(pady=10)
        lf=ttk.LabelFrame(c,text="Обозначения",padding=10); lf.pack(pady=15)
        for col,txt in [("red","Красная = Линейная"),("green","Зелёная = Полином2"),("blue","Синяя = Полином3")]:
            f=ttk.Frame(lf); f.pack(anchor=tk.W,pady=2)
            ttk.Label(f,text="●",foreground=col,font=("Calibri",12)).pack(side=tk.LEFT)
            ttk.Label(f,text=txt,font=("Calibri",12)).pack(side=tk.LEFT,padx=5)
        ef=ttk.LabelFrame(lf,text="Цвет ошибок",padding=5); ef.pack(anchor=tk.W,pady=5)
        for col,txt in [("green","Зелёный шрифт — ошибка < 1%"),("orange","Оранжевый шрифт — ошибка 1–5%"),("red","Красный шрифт — ошибка > 5%")]:
            ttk.Label(ef,text=f"● {txt}",foreground=col,font=("Calibri",11)).pack(anchor=tk.W)
    def lf(s):
        fp=filedialog.askopenfilename(filetypes=[("Все","*.txt;*.csv;*.xls;*.xlsx"),("CSV","*.csv"),("TXT","*.txt"),("Excel","*.xls;*.xlsx")])
        if not fp: return
        try: data=lex(fp) if fp.lower().endswith(('.xls','.xlsx')) else lcsv(fp)
        except Exception as e: messagebox.showerror("Ошибка",str(e)); return
        s.ra(*data)
    def lm(s):
        try: data=m_inp()
        except Exception as e:
            if str(e)!="Ввод отменён": messagebox.showerror("Ошибка",str(e))
            return
        s.ra(*data)
    def ra(s,x,y,xl):
        if len(np.unique(x))<2: messagebox.showerror("Ошибка","Недостаточно X"); return
        models={}
        for deg,nm in [(1,"Линейная"),(2,"Полином 2-й ст."),(3,"Полином 3-й ст.")]:
            if len(x)>=deg+1:
                t0=time.perf_counter(); coeffs=pf(x,y,deg); yp=pv(coeffs,x); r2v=r2(y,yp)
                models[nm]=(coeffs,r2v,time.perf_counter()-t0,yp)
            else: models[nm]=(None,None,None,None)
        fx=np.arange(np.max(x)+1,np.max(x)+4)
        ResultWindow(s.m,x,y,xl,models,fx)
class ResultWindow:
    def __init__(s,parent,x,y,xl,models,fx):
        w=tk.Toplevel(parent); w.title("Результаты"); w.geometry("1300x800")
        s.x, s.y, s.lb, s.models, s.fx = x,y,xl,models,fx
        m=ttk.Frame(w,padding=10); m.pack(fill=tk.BOTH,expand=True)
        ttk.Label(m,text="РЕЗУЛЬТАТЫ АППРОКСИМАЦИИ",font=("Times New Roman",14,"bold")).pack(pady=5)
        s._info(m); s._table(m); s._forecast(m)
        b=ttk.Frame(m); b.pack(pady=10)
        ttk.Button(b,text="📈 График",command=s.graph).pack(side=tk.LEFT,padx=5)
        ttk.Button(b,text="💾 Сохранить",command=s.save).pack(side=tk.LEFT,padx=5)
        leg=ttk.LabelFrame(m,text="Подсказка",padding=5); leg.pack(fill=tk.X,pady=5)
        ttk.Label(leg,text="Цвет ошибок: зелёный < 1%, оранжевый 1–5%, красный > 5%. Столбец «Лучшая» – модель с мин. ошибкой.",font=("Calibri",10)).pack(anchor=tk.W)
    def _fmt(s,coeffs,deg):
        if coeffs is None: return "—"
        if deg==1: return f"y = {coeffs[0]:.6f} + {coeffs[1]:.6f}·x"
        t=[f"{coeffs[0]:.6f}"]
        for i in range(1,deg+1):
            c=coeffs[i]; sg="+" if c>=0 else "-"; a=abs(c)
            t.append(f"{sg} {a:.6f}·x^{i}" if i>1 else f"{sg} {a:.6f}·x")
        return "y = "+" ".join(t)
    def _info(s,parent):
        f=ttk.LabelFrame(parent,text="Модели",padding=5); f.pack(fill=tk.X,pady=5)
        for nm,(coeffs,r2v,tm,_) in s.models.items():
            if coeffs is not None:
                deg=1 if nm=="Линейная" else (2 if "2" in nm else 3)
                txt=f"{nm}: {s._fmt(coeffs,deg)} | R² = {r2v:.6f} | {tm:.6f} c"
                ttk.Label(f,text=txt,font=("Consolas",10)).pack(anchor=tk.W)
    def _table(s,parent):
        f=ttk.LabelFrame(parent,text="Сравнение",padding=5); f.pack(fill=tk.BOTH,expand=True,pady=5)
        cols=["X","Y факт","Линейная","Ош.лин","Ош.лин%","Полином2","Ош.пол2","Ош.пол2%","Полином3","Ош.пол3","Ош.пол3%","Лучшая (мин. ошибка)"]
        tree=ttk.Treeview(f,columns=cols,show="headings",height=min(len(s.x),15))
        for c,w in zip(cols,[70,90,90,80,80,90,80,80,90,80,80,120]): tree.heading(c,text=c); tree.column(c,width=w,anchor=tk.CENTER)
        sb=ttk.Scrollbar(f,orient=tk.VERTICAL,command=tree.yview); tree.configure(yscrollcommand=sb.set)
        sb.pack(side=tk.RIGHT,fill=tk.Y); tree.pack(fill=tk.BOTH,expand=True)
        tree.tag_configure('le',foreground='green'); tree.tag_configure('me',foreground='orange')
        tree.tag_configure('he',foreground='red'); tree.tag_configure('br',background='#c8e6c9')
        s1,s2,s3=0.0,0.0,0.0
        for i in range(len(s.x)):
            row=[s.lb[i],f"{s.y[i]:.2f}"]; be=1e9; bn=None
            for nm in ["Линейная","Полином 2-й ст.","Полином 3-й ст."]:
                coeffs,_,_,yp=s.models[nm]
                if coeffs is not None:
                    pred=yp[i]; err=abs(s.y[i]-pred); perc=(err/abs(s.y[i]))*100 if s.y[i]!=0 else 0.0
                    if err<be: be=err; bn=nm
                    row.extend([f"{pred:.2f}",f"{err:.2f}",f"{perc:.1f}%"])
                    if nm=="Линейная": s1+=err
                    elif nm=="Полином 2-й ст.": s2+=err
                    else: s3+=err
                else: row.extend(["—","—","—"])
            tag='le' if be==0 else ('me' if be<0.5 else 'he')
            row.append({"Линейная":"Линейная","Полином 2-й ст.":"Полином2","Полином 3-й ст.":"Полином3"}.get(bn,"—"))
            tree.insert("",tk.END,values=row,tags=(tag,))
        tree.insert("",tk.END,values=["ИТОГО",""]+[f"{s1:.2f}","",""]+[f"{s2:.2f}","",""]+[f"{s3:.2f}","","",""],tags=('br',))
    def _forecast(s,parent):
        f=ttk.LabelFrame(parent,text="Прогноз на 3 шага",padding=5); f.pack(fill=tk.X,pady=5)
        ttk.Label(f,text="Модель с максимальным R² рекомендуется как самая надёжная:",font=("Calibri",11,"bold")).pack(anchor=tk.W,pady=5)
        for i,nx in enumerate(s.fx):
            vals={}
            for nm,(coeffs,_,_,_) in s.models.items():
                if coeffs is not None: vals[nm]=pv(coeffs,nx)
            if vals:
                best=max((m for m in s.models if s.models[m][0] is not None),key=lambda m:s.models[m][1])
                st=f"Шаг {i+1} (X={nx:.0f}): "+"   ".join([f"{nm}={v:.2f}" for nm,v in vals.items()])
                st+=f"   | Рекомендация: {best} (R²={s.models[best][1]:.4f})"
                ttk.Label(f,text=st,font=("Consolas",10)).pack(anchor=tk.W)
    def graph(s):
        plt.figure(figsize=(12,6)); plt.scatter(s.x,s.y,color='black',label='Данные',zorder=5)
        xp=np.linspace(min(s.x),max(s.x)+3,500)
        cols={'Линейная':'red','Полином 2-й ст.':'green','Полином 3-й ст.':'blue'}
        for nm,(coeffs,r2v,_,_) in s.models.items():
            if coeffs is not None:
                plt.plot(xp,pv(coeffs,xp),color=cols[nm],lw=2,label=f'{nm} (R²={r2v:.4f})')
                plt.scatter(s.fx,pv(coeffs,s.fx),color=cols[nm],marker='x',s=100)
                deg=1 if nm=="Линейная" else (2 if "2" in nm else 3)
                plt.text(0.02,0.98-list(s.models.keys()).index(nm)*0.05,s._fmt(coeffs,deg),
                         transform=plt.gca().transAxes,fontsize=8,verticalalignment='top',color=cols[nm])
        if s.lb and len(s.lb)==len(s.x):
            step=max(1,len(s.x)//20); pos=s.x[::step]; lbl=[s.lb[i] for i in range(0,len(s.x),step)]
            plt.xticks(pos,lbl,rotation=45,ha='right')
        plt.xlabel('X'); plt.ylabel('Y'); plt.title('Аппроксимация'); plt.legend()
        plt.grid(True,linestyle='--',alpha=0.6); plt.tight_layout(); plt.show()
    def save(s):
        fp=filedialog.asksaveasfilename(defaultextension=".txt",filetypes=[("Текст","*.txt")])
        if not fp: return
        with open(fp,'w',encoding='utf-8') as f:
            f.write("РЕЗУЛЬТАТЫ\n\n")
            for nm,(coeffs,r2v,tm,_) in s.models.items():
                if coeffs is not None:
                    deg=1 if nm=="Линейная" else (2 if "2" in nm else 3)
                    f.write(f"{nm}: {s._fmt(coeffs,deg)} | R²={r2v:.6f} | {tm:.6f}c\n")
            f.write("\nТаблица (X,Y,Лин,ОшЛ,ОшЛ%,П2,ОшП2,ОшП2%,П3,ОшП3,ОшП3%)\n")
            for i in range(len(s.x)):
                row=[s.lb[i],f"{s.y[i]:.2f}"]
                for nm in ["Линейная","Полином 2-й ст.","Полином 3-й ст."]:
                    coeffs,_,_,yp=s.models[nm]
                    if coeffs is not None:
                        pred=yp[i]; err=abs(s.y[i]-pred); perc=(err/abs(s.y[i]))*100 if s.y[i]!=0 else 0.0
                        row.extend([f"{pred:.2f}",f"{err:.2f}",f"{perc:.1f}%"])
                    else: row.extend(["—","—","—"])
                f.write("\t".join(row)+"\n")
            f.write("\nПрогноз:\n")
            for nx in s.fx:
                vals={}
                for nm,(coeffs,_,_,_) in s.models.items():
                    if coeffs is not None: vals[nm]=pv(coeffs,nx)
                f.write(f"X={nx:.0f}: "+" ".join([f"{nm}={v:.2f}" for nm,v in vals.items()])+"\n")
        messagebox.showinfo("Сохранено",f"Отчёт сохранён в {fp}")
if __name__=='__main__':
    root=tk.Tk(); app=App(root); root.mainloop()